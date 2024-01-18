# Importing the requests library to fetch web content
import requests
# Importing the BeautifulSoup module from bs4 for web scraping
import bs4
#importing openpyxl 
from openpyxl import Workbook, load_workbook 

wb=load_workbook('Cryptocurrency.xlsx')
# Setting the URL of the coinmarketcap dates to get prices
url = "https://coinmarketcap.com/"
url_zkf = "https://coinmarketcap.com/currencies/zkfair/"
url_matic = "https://coinmarketcap.com/currencies/polygon/"
url_atom = "https://coinmarketcap.com/currencies/cosmos/"
url_apt = "https://coinmarketcap.com/currencies/aptos/"
url_arb = "https://coinmarketcap.com/currencies/arbitrum/"
url_sui = "https://coinmarketcap.com/currencies/sui/"
# Using requests.get() to send a GET request to the specified URL and fetch its content
saturs = requests.get(url)
zkf_request = requests.get(url_zkf)
matic_request = requests.get(url_matic)
atom_request = requests.get(url_atom)
apt_request = requests.get(url_apt)
arb_request = requests.get(url_arb)
sui_request = requests.get(url_sui)

ws=wb.active
max_row=ws.max_row

for row in range(2,max_row+1):
    
    previus_price = ws['D' + str(row)].value
    ws['E' + str(row)].value = previus_price
    
wb.save('Cryptocurrency.xlsx')


# Checking if the request was successful (HTTP status code 200 means OK/successful)
if saturs.status_code == 200:
    #get price of cryptocurencies
    lapas_saturs = bs4.BeautifulSoup(saturs.content, 'html.parser')
    btc_price = lapas_saturs.find_all(href="/currencies/bitcoin/#markets")
    eth_price = lapas_saturs.find_all(href="/currencies/ethereum/#markets")
    bnb_price = lapas_saturs.find_all(href="/currencies/bnb/#markets")
    usdt_price = lapas_saturs.find_all(href="/currencies/tether/#markets")
    sol_price = lapas_saturs.find_all(href="/currencies/solana/#markets")
if zkf_request.status_code == 200:
    zkf_saturs = bs4.BeautifulSoup(zkf_request.content, 'html.parser')
    zkf_price = zkf_saturs.find_all(class_="sc-f70bb44c-0 jxpCgO base-text")
if matic_request.status_code == 200:
    matic_saturs = bs4.BeautifulSoup(matic_request.content, 'html.parser')
    matic_price = matic_saturs.find_all(class_="sc-f70bb44c-0 jxpCgO base-text")
if atom_request.status_code == 200:
    atom_saturs = bs4.BeautifulSoup(atom_request.content, 'html.parser')
    atom_price = atom_saturs.find_all(class_="sc-f70bb44c-0 jxpCgO base-text")
if apt_request.status_code == 200:
    apt_saturs = bs4.BeautifulSoup(apt_request.content, 'html.parser')
    apt_price = apt_saturs.find_all(class_="sc-f70bb44c-0 jxpCgO base-text")
if arb_request.status_code == 200:
    arb_saturs = bs4.BeautifulSoup(arb_request.content, 'html.parser')
    arb_price = arb_saturs.find_all(class_="sc-f70bb44c-0 jxpCgO base-text")
if sui_request.status_code == 200:
    sui_saturs = bs4.BeautifulSoup(sui_request.content, 'html.parser')
    sui_price = sui_saturs.find_all(class_="sc-f70bb44c-0 jxpCgO base-text")
    
#create list with prices and clear it from "$" and ","
tokens_prices = [btc_price[0].string,eth_price[0].string,bnb_price[0].string,matic_price[0].string,sol_price[0].string,atom_price[0].string,apt_price[0].string,arb_price[0].string,sui_price[0].string,usdt_price[0].string,zkf_price[0].string]
for i in range(0,11):
    tokens_prices[i] = str(tokens_prices[i]).replace(",","").replace("$","")
    tokens_prices[i] = float(tokens_prices[i])
    

ws=wb.active
max_row=ws.max_row

for row in range(2,max_row+1):
    
    ws['D' + str(row)].value = tokens_prices[row-2]

    
wb.save('Cryptocurrency.xlsx')
wb.close()  

    
